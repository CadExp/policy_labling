"""
@author: lxy
@email: linxy59@mail2.sysu.edu.cn
@date: 2021/9/29
@description: null
"""


def load_xlsx_as_dataframe(filename):
    import openpyxl
    from pandas import DataFrame
    book = openpyxl.load_workbook(filename)
    sheet = book.active
    data = []
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        row_data = [cell.value for cell in row]
        data.append(row_data)
    len(data)
    header = data[0]  # 第一行
    content_data = data[1:]
    dict_data = {header[i]: [row[i] for row in content_data] for i in range(len(header))}
    frame = DataFrame(dict_data)
    return frame


def dataframe_to_xlsx(frame):
    from openpyxl import Workbook
    new_book = Workbook()
    new_sheet = new_book.active
    cols_names = list(frame.columns)
    new_sheet.append(cols_names)
    for i in range(len(frame)):
        sheet_row = [frame.iloc[i][name] for name in cols_names]
        new_sheet.append(sheet_row)
    # new_book.save("sample.xlsx")
    return new_book
